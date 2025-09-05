from __future__ import annotations

import os
import io
import json
import re
import hashlib
from pathlib import Path
from dataclasses import dataclass, field
from datetime import datetime
from typing import Dict, Any, List, Optional
from urllib.parse import urlparse

import pandas as pd

# ========= Raíz del proyecto y carpeta results/ =========
# Permite forzar rutas por variables de entorno
_PROJECT_ROOT_ENV = os.getenv("PROJECT_ROOT")
_RESULTS_DIR_ENV = os.getenv("RESULTS_DIR")

PROJECT_ROOT = (
    _PROJECT_ROOT_ENV
    or os.path.abspath(os.path.join(os.path.dirname(__file__), "../.."))
)
DEFAULT_RESULTS_DIR = (
    _RESULTS_DIR_ENV
    or os.path.join(PROJECT_ROOT, "results")
)

JSON_SCHEMA_VERSION = "1.0.0"


def _ensure_results_dir(dir_path: Optional[str] = None) -> str:
    base = dir_path or DEFAULT_RESULTS_DIR
    abs_dir = os.path.abspath(os.path.expanduser(base))
    os.makedirs(abs_dir, exist_ok=True)
    return abs_dir


@dataclass
class ExcelResultsExporter:
    """Exporta resultados a Excel/JSON con el esquema requerido + tarjeta de decisión."""
    sheet_name: str = "RESULTADOS"
    show_domain_only: bool = False  # Si True, 'FUENTE' será solo el dominio
    date_input_formats: List[str] = field(default_factory=lambda: [
        "%d-%m-%Y", "%Y-%m-%d", "%Y/%m/%d", "%Y-%m-%dT%H:%M:%SZ", "%d/%m/%Y"
    ])

    EXPORT_COLUMNS: List[str] = field(default_factory=lambda: [
        "AÑO", "MES", "PAÍS", "INTENCIONALIDAD", "OBJETIVO", "COMPAÑÍA", "MODELO",
        "TIPO DE COMPAÑÍA", "DATOS/VIDA", "ACCESO", "SISTEMA", "ALCANCE",
        "REC. POR LA MARCA", "FUENTE", "DESCRIPCIÓN", "TITULAR",
    ])

    MESES_ES: Dict[int, str] = field(default_factory=lambda: {
        1: "ENERO", 2: "FEBRERO", 3: "MARZO", 4: "ABRIL", 5: "MAYO", 6: "JUNIO",
        7: "JULIO", 8: "AGOSTO", 9: "SEPTIEMBRE", 10: "OCTUBRE", 11: "NOVIEMBRE", 12: "DICIEMBRE"
    })

    TLD_TO_COUNTRY: Dict[str, str] = field(default_factory=lambda: {
        "es": "España", "mx": "México", "ar": "Argentina", "cl": "Chile", "co": "Colombia",
        "pe": "Perú", "uy": "Uruguay", "bo": "Bolivia", "py": "Paraguay", "ve": "Venezuela",
        "ec": "Ecuador", "cr": "Costa Rica", "pa": "Panamá", "hn": "Honduras",
        "ni": "Nicaragua", "sv": "El Salvador", "gt": "Guatemala",
        "do": "República Dominicana", "pr": "Puerto Rico",
    })

    # ----------------- Helpers de E/S ----------------- #
    def _abs(self, p: str) -> str:
        return str(Path(p).expanduser().resolve())

    def _assert_written(self, path: str) -> None:
        p = Path(path)
        if not p.exists():
            raise IOError(f"Archivo no encontrado tras escribir: {p}")
        if p.stat().st_size == 0:
            raise IOError(f"Archivo vacío tras escribir: {p}")

    # ----------------- PÚBLICOS: GUARDAR EN DISCO ----------------- #

    def save_single_to_disk(
        self,
        results_dict: dict,
        category: str,
        dir_path: Optional[str] = None,
        timestamp_str: Optional[str] = None,
    ) -> Optional[str]:
        target_dir = _ensure_results_dir(dir_path)
        df = self.build_export_df(results_dict)
        if df.empty:
            return None
        ts = timestamp_str or self._now_madrid_str()
        filename = f"results_{category}_{ts}.xlsx"
        path = os.path.join(target_dir, filename)
        with open(path, "wb") as f:
            f.write(self._to_excel_bytes(df, sheet_name=self.sheet_name))
        self._assert_written(path)
        return self._abs(path)

    def save_json_to_disk(
        self,
        results_dict: dict,
        category: str,
        dir_path: Optional[str] = None,
        timestamp_str: Optional[str] = None,
    ) -> Optional[str]:
        target_dir = _ensure_results_dir(dir_path)
        df = self.build_export_df(results_dict)
        if df.empty:
            return None
        data = df.to_dict(orient="records")
        ts = timestamp_str or self._now_madrid_str()
        filename = f"results_{category}_{ts}.json"
        path = os.path.join(target_dir, filename)
        with open(path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        self._assert_written(path)
        return self._abs(path)

    def save_multi_to_disk(
        self,
        results_by_category: dict,   # {"news": {...}, "papers": {...}, "vulnerabilities": {...}}
        dir_path: Optional[str] = None,
        timestamp_str: Optional[str] = None,
    ) -> Optional[str]:
        target_dir = _ensure_results_dir(dir_path)
        dfs_map = self._to_dfs_map_multi(results_by_category)
        if not dfs_map:
            return None
        ts = timestamp_str or self._now_madrid_str()
        filename = f"results_all_{ts}.xlsx"
        path = os.path.join(target_dir, filename)
        with open(path, "wb") as f:
            f.write(self._to_excel_bytes_multi(dfs_map))
        self._assert_written(path)
        return self._abs(path)

    def save_multi_json_to_disk(
        self,
        results_by_category: dict,
        dir_path: Optional[str] = None,
        timestamp_str: Optional[str] = None,
    ) -> Optional[str]:
        """Guarda un JSON combinado (hoja TODOS)."""
        target_dir = _ensure_results_dir(dir_path)
        dfs_map = self._to_dfs_map_multi(results_by_category)
        if not dfs_map or "TODOS" not in dfs_map:
            return None
        data = dfs_map["TODOS"].to_dict(orient="records")
        ts = timestamp_str or self._now_madrid_str()
        filename = f"results_all_{ts}.json"
        path = os.path.join(target_dir, filename)
        with open(path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        self._assert_written(path)
        return self._abs(path)

    # ----------------- CONSTRUCCIÓN DE DATAFRAMES ----------------- #

    def build_export_df(self, results_dict: Optional[Dict[str, Dict[str, Any]]]) -> pd.DataFrame:
        rows = []
        for _, item in (results_dict or {}).items():
            title = (item.get("Title") or "").strip()
            summary = (item.get("Summary") or "").strip()
            year = (item.get("Year") or "")
            date = (item.get("Date") or "")
            url = (item.get("URL") or "").strip()
            fuente = self._domain_from_url(url) if self.show_domain_only else url

            rows.append({
                "AÑO": year or "",
                "MES": self._parse_mes_from_date(date),
                "PAÍS": self._infer_pais_from_url(url),
                "INTENCIONALIDAD": "",
                "OBJETIVO": "",
                "COMPAÑÍA": item.get("Company", "") or "",
                "MODELO": item.get("Model", "") or "",
                "TIPO DE COMPAÑÍA": item.get("CompanyType", "") or "",
                "DATOS/VIDA": "",
                "ACCESO": "",
                "SISTEMA": "",
                "ALCANCE": "",
                "REC. POR LA MARCA": "",
                "FUENTE": fuente,
                "DESCRIPCIÓN": summary,
                "TITULAR": title,
            })

        df = pd.DataFrame(rows, columns=self.EXPORT_COLUMNS)
        for c in self.EXPORT_COLUMNS:
            if c in df.columns:
                df[c] = df[c].astype(str).where(df[c].notna(), "")
        return df

    def _to_dfs_map_multi(self, results_by_category: dict) -> dict:
        dfs = {}
        for cat, res in (results_by_category or {}).items():
            df = self.build_export_df(res)
            if not df.empty:
                dfs[cat] = df
        if not dfs:
            return {}
        df_all = pd.concat(dfs.values(), ignore_index=True)
        return {"TODOS": df_all, **dfs}

    # ----------------- BYTES: EXCEL ----------------- #

    def _to_excel_bytes(self, df: pd.DataFrame, sheet_name: str) -> bytes:
        buf = io.BytesIO()
        try:
            with pd.ExcelWriter(buf, engine="openpyxl") as w:
                df.to_excel(w, index=False, sheet_name=sheet_name)
                self._format_sheet(w, df, sheet_name)
        except Exception:
            buf = io.BytesIO()
            with pd.ExcelWriter(buf, engine="xlsxwriter") as w:
                df.to_excel(w, index=False, sheet_name=sheet_name)
                self._format_sheet(w, df, sheet_name)
        buf.seek(0)
        return buf.read()

    def _to_excel_bytes_multi(self, dfs_map: dict) -> bytes:
        buf = io.BytesIO()
        try:
            with pd.ExcelWriter(buf, engine="openpyxl") as w:
                for sheet, df in dfs_map.items():
                    df.to_excel(w, index=False, sheet_name=sheet)
                for sheet, df in dfs_map.items():
                    self._format_sheet(w, df, sheet)
        except Exception:
            buf = io.BytesIO()
            with pd.ExcelWriter(buf, engine="xlsxwriter") as w:
                for sheet, df in dfs_map.items():
                    df.to_excel(w, index=False, sheet_name=sheet)
                for sheet, df in dfs_map.items():
                    self._format_sheet(w, df, sheet)
        buf.seek(0)
        return buf.read()

    def _format_sheet(self, writer: pd.ExcelWriter, df: pd.DataFrame, sheet_name: str) -> None:
        """Ajuste de anchos y autofiltro. Soporta openpyxl/xlsxwriter."""
        try:
            ws = writer.sheets[sheet_name]
        except Exception:
            return

        cols = list(df.columns)
        # openpyxl
        try:
            from openpyxl.utils import get_column_letter
            for idx, col in enumerate(cols, 1):
                max_len = max(len(str(col)), *(len(str(x)) for x in df[col].head(200)))
                ws.column_dimensions[get_column_letter(idx)].width = min(max(12, max_len + 2), 60)
            ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{len(df) + 1}"
            return
        except Exception:
            pass
        # xlsxwriter
        try:
            for idx, col in enumerate(cols):
                max_len = max(len(str(col)), *(len(str(x)) for x in df[col].head(200)))
                width = min(max(12, max_len + 2), 60)
                ws.set_column(idx, idx, width)
            ws.autofilter(0, 0, len(df), len(cols) - 1)
        except Exception:
            pass

    # --- helpers tarjeta decisión --- #

    def _strip_port(self, host: str) -> str:
        return host.split(":")[0] if host else host

    def _normalize_domain(self, url: str) -> str:
        try:
            host = urlparse(url).netloc.lower()
            return self._strip_port(host)
        except Exception:
            return ""

    def _parse_sources(self, src) -> List[Dict[str, Optional[str]]]:
        """
        Convierte Source(s) como ["economictimes.indiatimes.com (GDELT)"] en:
        [{"domain": "economictimes.indiatimes.com", "via": "GDELT"}]
        """
        out: List[Dict[str, Optional[str]]] = []
        if not src:
            return out
        if isinstance(src, str):
            src = [src]
        for s in src:
            txt = str(s)
            m = re.match(r"^\s*([^()]+?)\s*(?:\(([^()]+)\))?\s*$", txt)
            if m:
                domain = self._strip_port((m.group(1) or "").strip().lower())
                via = (m.group(2) or "").strip() or None
                if domain:
                    out.append({"domain": domain, "via": via})
            else:
                out.append({"domain": self._strip_port(txt.strip().lower()), "via": None})
        return out

    def _aggregate_score(self, heur: Dict[str, Any]) -> float:
        """
        Score global 0–1 a partir de heurísticas.
        Normaliza automoción sobre 12 y incidente sobre 15 (cap al 1.0).
        """
        a = (heur.get("automotive") or {}).get("score")
        i = (heur.get("incident") or {}).get("score")
        parts: List[float] = []
        if isinstance(a, (int, float)):
            parts.append(min(1.0, max(0.0, a / 12.0)))
        if isinstance(i, (int, float)):
            parts.append(min(1.0, max(0.0, i / 15.0)))
        return sum(parts) / len(parts) if parts else 0.0

    def _certeza_from_score(self, s: float) -> str:
        if s >= 0.7:
            return "alta"
        if s >= 0.4:
            return "media"
        return "baja"

    def _bool_automotive_pass(self, heur: Dict[str, Any]) -> bool:
        h = heur.get("automotive") or {}
        sc = h.get("score") or 0
        hits = h.get("hits") or {}
        has_brand_or_term = bool((hits.get("brands") or []) or (hits.get("automotive_terms") or []))
        return bool(sc) or has_brand_or_term

    def _bool_incident_pass(self, heur: Dict[str, Any]) -> bool:
        h = heur.get("incident") or {}
        sc = h.get("score") or 0
        rs = " ".join(h.get("reasons") or []).lower()
        keys = ("attack_confirmed", "data_theft", "breach", "ransom", "shutdown", "vuln_found", "vulnerability", "cyberattack")
        any_key = any(k in rs for k in keys)
        return bool(sc) or any_key

    def _decide(self, heur: Dict[str, Any], ml_levels: List[Dict[str, Any]]):
        auto_pass = self._bool_automotive_pass(heur)
        inc_pass = self._bool_incident_pass(heur)
        ai_used = any(lvl.get("accepted") is True for lvl in ml_levels)
        ai_decision = "abstention" if not ai_used else ("positive" if any(l.get("accepted") for l in ml_levels) else "negative")
        decision = "keep" if (auto_pass and inc_pass) else "drop"
        return decision, auto_pass, inc_pass, ai_decision

    def _human_reasons(self, heur: Dict[str, Any]) -> List[str]:
        reasons: List[str] = []
        hits = (heur.get("automotive") or {}).get("hits") or {}
        inc = (heur.get("incident") or {})
        if hits.get("brands"):
            reasons.append(f"Marca detectada: {', '.join(hits['brands'])}")
        if hits.get("attack_terms"):
            reasons.append(f"Menciona: {', '.join(hits['attack_terms'])}")
        if hits.get("outcomes"):
            reasons.append(f"Resultado: {', '.join(hits['outcomes'])}")
        if hits.get("attack_vectors"):
            reasons.append(f"Vector: {', '.join(hits['attack_vectors'])}")
        if inc.get("reasons"):
            legibles = [re.sub(r"^\+?-?\d+\s*", "", r) for r in (inc.get("reasons") or [])]
            reasons.extend(legibles)
        # dedup + top-5
        seen = {}
        out = []
        for r in reasons:
            if r and r not in seen:
                seen[r] = 1
                out.append(r)
            if len(out) >= 5:
                break
        return out

    def _category_label(self, input_cat: str, auto_pass: bool, inc_pass: bool) -> str:
        if not auto_pass:
            return "no automoción"
        if not inc_pass:
            return "no incidente"
        mapping = {
            "news": "noticia automoción",
            "papers": "paper automoción",
            "vulnerabilities": "CVE automoción"
        }
        return mapping.get(input_cat, "contenido automoción")

    def _build_trace(self,
                     heur: Dict[str, Any],
                     ai_decision: str,
                     rules_version: Optional[str],
                     run_id: str,
                     export_ts_iso: str) -> Dict[str, Any]:
        auto_reasons = self._human_reasons({"automotive": heur.get("automotive") or {}, "incident": {}})[:3]
        inc_reasons = [re.sub(r"^\+?-?\d+\s*", "", r) for r in ((heur.get("incident") or {}).get("reasons") or [])][:3]
        return {
            "order": ["automotive_filter", "incident_filter", "ai_zeroshot"],
            "automotive_filter": {
                "decision": "pass" if self._bool_automotive_pass(heur) else "fail",
                "reasons": auto_reasons
            },
            "incident_filter": {
                "decision": "pass" if self._bool_incident_pass(heur) else "fail",
                "reasons": inc_reasons
            },
            "ai_zeroshot": {
                "decision": ai_decision,
                "label": "NO_LABEL" if ai_decision == "abstention" else None,
                "confidence": None,
                "model_version": None
            },
            "rules_version": rules_version or "rules@v1",
            "pipeline": {
                "run_id": run_id,
                "timestamp": export_ts_iso
            }
        }

    # ----------------- PRIVADOS: PARSERS & FECHA ----------------- #

    def _parse_mes_from_date(self, date_str: str) -> str:
        if not date_str:
            return ""
        for fmt in self.date_input_formats:
            try:
                dt = datetime.strptime(date_str, fmt)
                return self.MESES_ES.get(dt.month, "")
            except Exception:
                continue
        return ""

    def _infer_pais_from_url(self, url: str) -> str:
        try:
            netloc = urlparse(url).netloc.lower()
            tld = netloc.split(".")[-1]
            return self.TLD_TO_COUNTRY.get(tld, "")
        except Exception:
            return ""

    def _domain_from_url(self, url: str) -> str:
        try:
            return urlparse(url).netloc.lower()
        except Exception:
            return url

    def _now_madrid_str(self) -> str:
        """Fecha/hora Europe/Madrid → 'DD-MM-YYYY_HH-MM'."""
        try:
            from zoneinfo import ZoneInfo
            dt = datetime.now(ZoneInfo("Europe/Madrid"))
        except Exception:
            dt = datetime.now()
        return dt.strftime("%d-%m-%Y_%H-%M")

    def _normalize_date_iso(self, date_str: str) -> str:
        """Intenta convertir fechas tipo 'DD-MM-YYYY', 'YYYY/MM/DD', etc. a 'YYYY-MM-DD'."""
        if not date_str:
            return ""
        fmts = self.date_input_formats if isinstance(self.date_input_formats, list) else [
            "%d-%m-%Y", "%Y-%m-%d", "%Y/%m/%d", "%d/%m/%Y"
        ]
        for fmt in fmts:
            try:
                dt = datetime.strptime(date_str, fmt)
                return dt.strftime("%Y-%m-%d")
            except Exception:
                continue
        m = re.match(r"^\d{4}-\d{2}-\d{2}$", (date_str or "").strip())
        return date_str if m else ""

    def _collect_ml_levels(self, item: Dict[str, Any]) -> list:
        """
        Agrupa los campos ML por nivel:
        Soporta claves 'Label_level_1' y 'Label_1' (y análogas para Score/Accepted/...).
        """
        levels = set()
        for k in item.keys():
            m = re.match(r"Label_(?:level_)?(\d+)$", k)
            if m:
                levels.add(int(m.group(1)))
        out = []
        for lvl in sorted(levels):
            suffix_a = f"level_{lvl}"
            suffix_b = f"{lvl}"
            out.append({
                "level": lvl,
                "label": item.get(f"Label_{suffix_a}") or item.get(f"Label_{suffix_b}"),
                "score": item.get(f"Score_{suffix_a}") or item.get(f"Score_{suffix_b}"),
                "accepted": item.get(f"Accepted_{suffix_a}") or item.get(f"Accepted_{suffix_b}"),
                "votes": item.get(f"Votes_{suffix_a}") or item.get(f"Votes_{suffix_b}"),
                "margin": item.get(f"Margin_{suffix_a}") or item.get(f"Margin_{suffix_b}"),
                "entropy": item.get(f"Entropy_{suffix_a}") or item.get(f"Entropy_{suffix_b}")
            })
        return out

    def _content_hash(self, payload: Dict[str, Any]) -> str:
        """Hash estable del contenido base (Title+Summary+URL) para trazabilidad ligera."""
        base = {
            "title": payload.get("Title") or "",
            "summary": payload.get("Summary") or "",
            "url": payload.get("URL") or "",
        }
        txt = json.dumps(base, ensure_ascii=False, sort_keys=True)
        return "sha256:" + hashlib.sha256(txt.encode("utf-8")).hexdigest()

    def _to_enriched_record(self, item: Dict[str, Any], category: str, export_ts_iso: str) -> Dict[str, Any]:
        url = (item.get("URL") or "").strip()
        domain = self._normalize_domain(url) if url else ""
        published_date_iso = self._normalize_date_iso(item.get("Date"))

        heur = {
            "automotive": {
                "score": item.get("Heur_Score"),
                "tags": item.get("Heur_Tags"),
                "hits": item.get("Heur_Hits"),
            },
            "incident": {
                "score": item.get("IncidentScore"),
                "reasons": item.get("IncidentReasons"),
                "category": item.get("IncidentCategory"),
            }
        }

        ml_levels = self._collect_ml_levels(item)

        # --- tarjeta de decisión (base heurística) --- #
        sources = self._parse_sources(item.get("Source") or item.get("Sources") or [])
        domains = list({s["domain"] for s in sources if s.get("domain")}) or ([domain] if domain else [])
        sources_count = len(sources)

        score_global = self._aggregate_score(heur)
        certeza = self._certeza_from_score(score_global)
        decision, auto_pass, inc_pass, ai_decision = self._decide(heur, ml_levels)
        reasons_human = self._human_reasons(heur)
        category_label = self._category_label(category, auto_pass, inc_pass)

        run_id = item.get("RunId") or self._content_hash(item)
        rules_version = item.get("RulesVersion") or "rules@v1"

        signals_negative = []
        try:
            negs = ((heur.get("automotive") or {}).get("tags") or {}).get("negatives") or []
            if isinstance(negs, list):
                signals_negative = negs
        except Exception:
            pass

        tw_from = item.get("TimeFrom") or None
        tw_to = item.get("TimeTo") or None
        time_window = {"from": tw_from, "to": tw_to} if (tw_from or tw_to) else None
        consolidation = {"duplicates_merged": item.get("DupMerged", 0)}

        # --- OVERRIDES del FilterEngine (si existen) ---
        if item.get("Decision") in ("keep", "drop"):
            decision = item["Decision"]
        reasons_human = item.get("DecisionReasons") or reasons_human
        category_label = item.get("CategoryLabel") or category_label
        decision_gate = item.get("DecisionGate")  # "automotive_filter" | "incident_filter" | "ai_zeroshot" | "final"

        trace = self._build_trace(heur, ai_decision, rules_version, run_id, export_ts_iso)
        trace["decision_source"] = decision_gate or "heuristics"
        trace["final_decision"] = decision

        rec = {
            "schema_version": JSON_SCHEMA_VERSION,

            # --- cabecera "humana" de la tarjeta ---
            "category_input": category,
            "category_label": category_label,
            "language": item.get("Language"),
            "published_date": published_date_iso or (item.get("Date") or ""),
            "title_normalized": (item.get("Title") or "").strip(),
            "canonical_url": url,

            "sources": sources,
            "domains": domains,
            "sources_count": sources_count,

            # --- decisión de filtrado (contrato 5.4) ---
            "decision": decision,                             # keep | drop
            "score_global": round(float(score_global), 4),
            "certeza": certeza,                               # baja | media | alta
            "reasons": reasons_human[:5],
            "trace": trace,

            # --- metadatos de auditoría ---
            "time_window": time_window,
            "consolidation": consolidation,
            "signals_negative": signals_negative,
            "human_review": {"status": "pendiente", "comment": None},

            # --- payload original ---
            "category": category,
            "item_id": item.get("ID") or url or self._content_hash(item),
            "raw": {
                "title": item.get("Title") or "",
                "summary": item.get("Summary") or "",
                "url": url,
                "language": item.get("Language"),
                "published_year": item.get("Year"),
                "published_date": published_date_iso or (item.get("Date") or "")
            },
            "normalized": {
                "domain": domain
            },
            "heuristics": heur,
            "ml": ml_levels,
            "timestamps": {
                "search_ts": item.get("SearchTimestamp"),
                "export_ts": export_ts_iso
            },
            "metrics": item.get("Perf") or {}
        }

        return rec

    def save_json_enriched_to_disk(
        self,
        results_dict: Dict[str, Dict[str, Any]],
        category: str,
        dir_path: Optional[str] = None,
        timestamp_str: Optional[str] = None,
        ndjson: bool = True
    ) -> Optional[str]:
        """
        Guarda JSON enriquecido por ítem:
        - Por defecto NDJSON (.jsonl) para análisis a escala.
        - Si ndjson=False, guarda lista JSON tradicional.
        """
        target_dir = _ensure_results_dir(dir_path)
        if not results_dict:
            return None

        ts = timestamp_str or self._now_madrid_str()
        export_ts_iso = datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ")

        enriched = [
            self._to_enriched_record(item, category=category, export_ts_iso=export_ts_iso)
            for _, item in results_dict.items()
        ]
        if not enriched:
            return None

        if ndjson:
            filename = f"results_enriched_{category}_{ts}.jsonl"
            path = os.path.join(target_dir, filename)
            with open(path, "w", encoding="utf-8") as f:
                for rec in enriched:
                    f.write(json.dumps(rec, ensure_ascii=False) + "\n")
            self._assert_written(path)
            return self._abs(path)
        else:
            filename = f"results_enriched_{category}_{ts}.json"
            path = os.path.join(target_dir, filename)
            with open(path, "w", encoding="utf-8") as f:
                json.dump(enriched, f, ensure_ascii=False, indent=2)
            self._assert_written(path)
            return self._abs(path)

    def save_multi_json_enriched_to_disk(
        self,
        results_by_category: Dict[str, Dict[str, Dict[str, Any]]],
        dir_path: Optional[str] = None,
        timestamp_str: Optional[str] = None,
        ndjson: bool = True
    ) -> Optional[str]:
        """
        Combina varias categorías en un único JSON enriquecido.
        - NDJSON recomendado para ingestión analítica.
        """
        target_dir = _ensure_results_dir(dir_path)
        if not results_by_category:
            return None

        ts = timestamp_str or self._now_madrid_str()
        export_ts_iso = datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ")

        enriched_all = []
        for cat, res in (results_by_category or {}).items():
            for _, item in (res or {}).items():
                enriched_all.append(self._to_enriched_record(item, category=cat, export_ts_iso=export_ts_iso))

        if not enriched_all:
            return None

        if ndjson:
            filename = f"results_enriched_all_{ts}.jsonl"
            path = os.path.join(target_dir, filename)
            with open(path, "w", encoding="utf-8") as f:
                for rec in enriched_all:
                    f.write(json.dumps(rec, ensure_ascii=False) + "\n")
            self._assert_written(path)
            return self._abs(path)
        else:
            filename = f"results_enriched_all_{ts}.json"
            path = os.path.join(target_dir, filename)
            with open(path, "w", encoding="utf-8") as f:
                json.dump(enriched_all, f, ensure_ascii=False, indent=2)
            self._assert_written(path)
            return self._abs(path)
